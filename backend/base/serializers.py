from rest_framework import serializers
from rest_framework.response import Response
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from io import BytesIO
from typing import List, Dict, Any


class BaseSerializer(serializers.ModelSerializer):
  """
  Базовый сериализатор с поддержкой экспорта полей.
  """

  def __init__(self, *args: Any, **kwargs: Any) -> None:
    """
    Инициализация сериализатора и установка заголовков для экспорта.
    """
    super().__init__(*args, **kwargs)
    self.headers = self.get_export_fields()

  def get_export_fields(self) -> List[Dict[str, str]]:
    """Получение полей для экспорта.

    Returns:
        Список словарей с именем и verbose_name полей.
    """

    if not hasattr(self, 'Meta') or not hasattr(self.Meta, 'export_fields'):
      return []

    if '__all__' in self.Meta.export_fields:
      return self.__get_all_export_fields()
    else:
      return self._get_specific_export_fields(self.Meta.export_fields)

  def __get_all_export_fields(self) -> List[Dict[str, str]]:
    """
    Получение всех полей для экспорта.

    Returns:
    Список словарей с именем и лейблом полей.
    """
    headers = []
    for field_name, field in self.get_fields().items():
      headers.append({'name': field_name, 'verbose_name': field.label or field_name})
    return headers

  def _get_specific_export_fields(
    self, export_fields: List[str]
  ) -> List[Dict[str, str]]:
    """
    Получение конкретных полей для экспорта

    Args:
        export_fields (List[str]): Список имен полей для экспорта

    Returns:
        List[Dict[str, str]]: Список словарей с именем и лейблом полей
    """

    headers = []
    for field_name in export_fields:
      field = self.fields.get(field_name)
      if field:
        headers.append({'name': field_name, 'verbose_name': field.label or field_name})
    return headers

  def export_to_xlsx(self, queryset, filename: str = 'export.xlsx') -> Response:
    """Экспорт данных в файл XLSX

    Args:
        queryset (_type_): Данные, которые необходимо экспортировать
        filename (str, optional): Название файла. По дефолту 'export.xlsx'.

    Returns:
        Response: Ответ с экспортированным файлом XLSX
    """

    serializer = self.__class__(queryset, many=True)
    serialized_data = serializer.data

    result_wb = Workbook()
    result_wb.remove(result_wb.active)
    result_ws = result_wb.create_sheet(title='Выгрузка')

    # Заголовки
    result_ws.append([header['verbose_name'] for header in self.headers])

    # Данные
    for row_num, row in enumerate(serialized_data, start=2):
      for col_num, col in enumerate(self.headers, start=1):
        field = col['name']
        result_ws.cell(row=row_num, column=col_num).value = row.get(field, '')

    # Таблица
    result_table = Table(
      displayName='Table_Result',
      ref=f'A1:{get_column_letter(len(self.headers))}{len(serialized_data)+1}',
    )
    style = TableStyleInfo(
      name='TableStyleMedium2',
      showFirstColumn=False,
      showLastColumn=False,
      showRowStripes=True,
      showColumnStripes=False,
    )
    result_table.tableStyleInfo = style
    result_ws.add_table(result_table)

    # Создание ответа
    output_wb = BytesIO()
    result_wb.save(output_wb)
    result_wb.save('dasdada.xlsx')
    output_wb.seek(0)

    response = HttpResponse(
      content=output_wb.getvalue(),
      content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="asda.xlsx"'

    return response
